"""리뷰 CRUD API 라우터."""

import os
import sqlite3
from io import BytesIO
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook

from app import config
from app.database import get_db_dependency
from app.models import (
    ExcelError,
    ExcelUploadResult,
    ImageResponse,
    ReviewCreate,
    ReviewListResponse,
    ReviewResponse,
    ReviewUpdate,
    StatsResponse,
)

router = APIRouter(prefix="/api", tags=["reviews"])


# ---------------------------------------------------------------------------
# Helper functions
# ---------------------------------------------------------------------------


def _get_images(review_id: int, db: sqlite3.Connection) -> list[ImageResponse]:
    """주어진 리뷰의 이미지 목록을 반환한다."""
    cursor = db.execute(
        "SELECT id, review_id, file_path, original_name, file_size, created_at "
        "FROM review_images WHERE review_id = ? ORDER BY id",
        (review_id,),
    )
    rows = cursor.fetchall()
    return [
        ImageResponse(
            id=row["id"],
            review_id=row["review_id"],
            file_path=row["file_path"],
            original_name=row["original_name"],
            file_size=row["file_size"],
            created_at=str(row["created_at"]),
        )
        for row in rows
    ]


def _row_to_review(row: sqlite3.Row, db: sqlite3.Connection) -> ReviewResponse:
    """sqlite3.Row를 ReviewResponse로 변환한다 (이미지 포함)."""
    images = _get_images(row["id"], db)
    return ReviewResponse(
        id=row["id"],
        product_no=row["product_no"],
        product_name=row["product_name"] or "",
        author=row["author"],
        rating=row["rating"],
        title=row["title"] or "",
        content=row["content"],
        is_visible=bool(row["is_visible"]),
        display_order=row["display_order"] or 0,
        created_at=str(row["created_at"]),
        updated_at=str(row["updated_at"]),
        images=images,
    )


# ---------------------------------------------------------------------------
# 1. GET /api/reviews  -- 리뷰 목록 (페이징, 필터)
# ---------------------------------------------------------------------------


@router.get("/reviews", response_model=ReviewListResponse)
def list_reviews(
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100),
    product_no: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    db: sqlite3.Connection = Depends(get_db_dependency),
) -> ReviewListResponse:
    """리뷰 목록을 페이징 및 필터와 함께 반환한다."""
    conditions: list[str] = []
    params: list[object] = []

    if product_no:
        conditions.append("product_no = ?")
        params.append(product_no)

    if search:
        conditions.append(
            "(author LIKE ? OR title LIKE ? OR content LIKE ?)"
        )
        like_val = f"%{search}%"
        params.extend([like_val, like_val, like_val])

    where_clause = ""
    if conditions:
        where_clause = "WHERE " + " AND ".join(conditions)

    # 전체 건수
    count_sql = f"SELECT COUNT(*) as cnt FROM reviews {where_clause}"
    total: int = db.execute(count_sql, params).fetchone()["cnt"]

    # 페이징 데이터
    offset = (page - 1) * per_page
    data_sql = (
        f"SELECT * FROM reviews {where_clause} "
        f"ORDER BY created_at DESC LIMIT ? OFFSET ?"
    )
    rows = db.execute(data_sql, [*params, per_page, offset]).fetchall()

    items = [_row_to_review(row, db) for row in rows]
    return ReviewListResponse(
        items=items,
        total=total,
        page=page,
        per_page=per_page,
    )


# ---------------------------------------------------------------------------
# 9. GET /api/stats  -- 대시보드 통계
# ---------------------------------------------------------------------------


@router.get("/stats", response_model=StatsResponse)
def get_stats(
    db: sqlite3.Connection = Depends(get_db_dependency),
) -> StatsResponse:
    """대시보드 통계를 반환한다."""
    row = db.execute(
        "SELECT "
        "  COUNT(*) as total_reviews, "
        "  SUM(CASE WHEN is_visible = 1 THEN 1 ELSE 0 END) as visible_reviews, "
        "  COALESCE(AVG(rating), 0) as average_rating, "
        "  COUNT(DISTINCT product_no) as total_products "
        "FROM reviews"
    ).fetchone()

    return StatsResponse(
        total_reviews=row["total_reviews"],
        visible_reviews=row["visible_reviews"] or 0,
        average_rating=round(float(row["average_rating"]), 1),
        total_products=row["total_products"],
    )


# ---------------------------------------------------------------------------
# 7. POST /api/reviews/excel-upload  -- 엑셀 일괄 업로드
#    (review_id 경로 매개변수보다 먼저 등록해야 경로 충돌 방지)
# ---------------------------------------------------------------------------


@router.post("/reviews/excel-upload", response_model=ExcelUploadResult)
async def excel_upload(
    file: UploadFile,
    db: sqlite3.Connection = Depends(get_db_dependency),
) -> ExcelUploadResult:
    """엑셀 파일(.xlsx)로 리뷰를 일괄 등록한다."""
    if not file.filename or not file.filename.endswith(".xlsx"):
        raise HTTPException(
            status_code=400,
            detail="xlsx 파일만 업로드할 수 있습니다.",
        )

    contents = await file.read()
    try:
        wb = load_workbook(filename=BytesIO(contents), read_only=True)
    except Exception:
        raise HTTPException(
            status_code=400,
            detail="엑셀 파일을 읽을 수 없습니다.",
        )

    ws = wb.active
    if ws is None:
        raise HTTPException(
            status_code=400,
            detail="엑셀 파일에 활성 시트가 없습니다.",
        )

    rows_iter = ws.iter_rows(min_row=2, values_only=True)  # 헤더 건너뜀
    success_count = 0
    fail_count = 0
    errors: list[ExcelError] = []

    for idx, row_data in enumerate(rows_iter, start=2):
        # 빈 행 건너뛰기
        if not row_data or all(cell is None for cell in row_data):
            continue

        # 최소 6열이어야 함 (부족하면 None으로 채움)
        padded = list(row_data) + [None] * (6 - len(row_data))
        product_no_val = padded[0]
        product_name_val = padded[1]
        author_val = padded[2]
        rating_val = padded[3]
        title_val = padded[4]
        content_val = padded[5]

        # 유효성 검사
        row_errors: list[str] = []

        if not product_no_val:
            row_errors.append("상품번호는 필수입니다")
        if not author_val:
            row_errors.append("작성자명은 필수입니다")
        if not content_val:
            row_errors.append("리뷰내용은 필수입니다")

        # 별점 검사
        try:
            rating_int = int(rating_val)  # type: ignore[arg-type]
            if rating_int < 1 or rating_int > 5:
                row_errors.append("별점은 1~5 사이 정수여야 합니다")
        except (TypeError, ValueError):
            row_errors.append("별점은 1~5 사이 정수여야 합니다")
            rating_int = 0

        if row_errors:
            fail_count += 1
            errors.append(
                ExcelError(row=idx, message="; ".join(row_errors))
            )
            continue

        db.execute(
            "INSERT INTO reviews (product_no, product_name, author, rating, title, content) "
            "VALUES (?, ?, ?, ?, ?, ?)",
            (
                str(product_no_val).strip(),
                str(product_name_val).strip() if product_name_val else "",
                str(author_val).strip(),
                rating_int,
                str(title_val).strip() if title_val else "",
                str(content_val).strip(),
            ),
        )
        success_count += 1

    wb.close()

    return ExcelUploadResult(
        success_count=success_count,
        fail_count=fail_count,
        errors=errors,
    )


# ---------------------------------------------------------------------------
# 8. GET /api/reviews/excel-template  -- 엑셀 템플릿 다운로드
# ---------------------------------------------------------------------------


@router.get("/reviews/excel-template")
def excel_template() -> StreamingResponse:
    """리뷰 일괄 등록용 엑셀 템플릿을 다운로드한다."""
    wb = Workbook()
    ws = wb.active
    if ws is not None:
        ws.title = "리뷰"
        ws.append(["상품번호", "상품명", "작성자명", "별점(1~5)", "리뷰제목", "리뷰내용"])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    wb.close()

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="staff_review_template.xlsx"'
        },
    )


# ---------------------------------------------------------------------------
# 2. POST /api/reviews  -- 리뷰 생성
# ---------------------------------------------------------------------------


@router.post("/reviews", response_model=ReviewResponse, status_code=201)
def create_review(
    body: ReviewCreate,
    db: sqlite3.Connection = Depends(get_db_dependency),
) -> ReviewResponse:
    """새 리뷰를 생성한다."""
    cursor = db.execute(
        "INSERT INTO reviews "
        "(product_no, product_name, author, rating, title, content, is_visible, display_order) "
        "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
        (
            body.product_no,
            body.product_name,
            body.author,
            body.rating,
            body.title,
            body.content,
            int(body.is_visible),
            body.display_order,
        ),
    )
    review_id = cursor.lastrowid
    row = db.execute("SELECT * FROM reviews WHERE id = ?", (review_id,)).fetchone()
    return _row_to_review(row, db)


# ---------------------------------------------------------------------------
# 3. GET /api/reviews/{review_id}  -- 리뷰 상세
# ---------------------------------------------------------------------------


@router.get("/reviews/{review_id}", response_model=ReviewResponse)
def get_review(
    review_id: int,
    db: sqlite3.Connection = Depends(get_db_dependency),
) -> ReviewResponse:
    """단건 리뷰를 반환한다."""
    row = db.execute("SELECT * FROM reviews WHERE id = ?", (review_id,)).fetchone()
    if row is None:
        raise HTTPException(status_code=404, detail="리뷰를 찾을 수 없습니다.")
    return _row_to_review(row, db)


# ---------------------------------------------------------------------------
# 4. PUT /api/reviews/{review_id}  -- 리뷰 수정
# ---------------------------------------------------------------------------


@router.put("/reviews/{review_id}", response_model=ReviewResponse)
def update_review(
    review_id: int,
    body: ReviewUpdate,
    db: sqlite3.Connection = Depends(get_db_dependency),
) -> ReviewResponse:
    """기존 리뷰를 수정한다. 전달된 필드만 업데이트한다."""
    existing = db.execute(
        "SELECT * FROM reviews WHERE id = ?", (review_id,)
    ).fetchone()
    if existing is None:
        raise HTTPException(status_code=404, detail="리뷰를 찾을 수 없습니다.")

    update_data = body.model_dump(exclude_unset=True)
    if not update_data:
        return _row_to_review(existing, db)

    set_clauses: list[str] = []
    values: list[object] = []
    for field, value in update_data.items():
        if field == "is_visible":
            value = int(value)  # type: ignore[assignment]
        set_clauses.append(f"{field} = ?")
        values.append(value)

    set_clauses.append("updated_at = CURRENT_TIMESTAMP")
    values.append(review_id)

    sql = f"UPDATE reviews SET {', '.join(set_clauses)} WHERE id = ?"
    db.execute(sql, values)

    row = db.execute("SELECT * FROM reviews WHERE id = ?", (review_id,)).fetchone()
    return _row_to_review(row, db)


# ---------------------------------------------------------------------------
# 5. DELETE /api/reviews/{review_id}  -- 리뷰 삭제
# ---------------------------------------------------------------------------


@router.delete("/reviews/{review_id}")
def delete_review(
    review_id: int,
    db: sqlite3.Connection = Depends(get_db_dependency),
) -> dict[str, str]:
    """리뷰를 삭제한다. 연결된 이미지도 DB 및 파일시스템에서 함께 삭제한다."""
    existing = db.execute(
        "SELECT * FROM reviews WHERE id = ?", (review_id,)
    ).fetchone()
    if existing is None:
        raise HTTPException(status_code=404, detail="리뷰를 찾을 수 없습니다.")

    # 연결된 이미지 파일 삭제
    images = db.execute(
        "SELECT file_path FROM review_images WHERE review_id = ?", (review_id,)
    ).fetchall()
    for img in images:
        file_path = img["file_path"]
        full_path = os.path.join(config.UPLOAD_DIR, file_path)
        if os.path.isfile(full_path):
            os.remove(full_path)

    # DB 레코드 삭제 (CASCADE로 이미지 레코드도 삭제됨)
    db.execute("DELETE FROM reviews WHERE id = ?", (review_id,))

    return {"detail": "삭제되었습니다"}


# ---------------------------------------------------------------------------
# 6. PATCH /api/reviews/{review_id}/visibility  -- 노출 상태 토글
# ---------------------------------------------------------------------------


@router.patch("/reviews/{review_id}/visibility", response_model=ReviewResponse)
def toggle_visibility(
    review_id: int,
    body: dict,
    db: sqlite3.Connection = Depends(get_db_dependency),
) -> ReviewResponse:
    """리뷰의 노출 상태를 변경한다."""
    if "is_visible" not in body:
        raise HTTPException(
            status_code=400, detail="is_visible 필드가 필요합니다."
        )

    existing = db.execute(
        "SELECT * FROM reviews WHERE id = ?", (review_id,)
    ).fetchone()
    if existing is None:
        raise HTTPException(status_code=404, detail="리뷰를 찾을 수 없습니다.")

    is_visible = int(bool(body["is_visible"]))
    db.execute(
        "UPDATE reviews SET is_visible = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
        (is_visible, review_id),
    )

    row = db.execute("SELECT * FROM reviews WHERE id = ?", (review_id,)).fetchone()
    return _row_to_review(row, db)
